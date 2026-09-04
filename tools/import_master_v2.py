"""Turn MtG Deck Master into the JSON the viewer reads.

Two workbooks, and which one is authoritative matters. The v2 workbook has
never been opened in Excel, so every formula column comes back empty: Status,
Qty, Own, Extra (Bench), Buy Count and $ To Buy are blank in the file. The Flat
workbook carries the same rows with those columns already resolved, so it is
the source of truth when it is present.

The formula transcription stays anyway, and runs as a cross-check: each formula
is quoted above the code implementing it, and a disagreement with the flat sheet
stops the import rather than shipping a number nobody can account for. Today the
two agree on all 579 rows across all six derived columns.

Read-only. Neither workbook is ever written back.
"""
import json, os, re, sys, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

# D1..D6, in the order the Read Me sheet lists them. The short label is what the
# rest of the workbook uses; the full name is what Scryfall answers to.
DECKS = [
    ("D1", "Quintorius", "Quintorius, Loremaster"),
    ("D2", "Chulane",    "Chulane, Teller of Tales"),
    ("D3", "Atraxa",     "Atraxa, Praetors' Voice"),
    ("D4", "Betor",      "Betor, Ancestor's Voice"),
    ("D5", "Shadrix",    "Shadrix Silverquill"),
    ("D6", "Purphoros",  "Purphoros, God of the Forge"),
]
IDS = [d[0] for d in DECKS]


def num(v):
    """A blank cell and a zero are different things on this sheet, but for the
    arithmetic both count as nothing."""
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def blank(v):
    return v is None or str(v).strip() == ""


def status_of(row):
    """Column B, transcribed from the sheet.

    E=Qty  F=Own  J=Ordered  D=Bracket
    M:R = the six deck target columns      T:Y = the six deck actual columns
    """
    E, F, J, D = row["qty"], row["own"], row["ordered"], row["bracket"]
    deck_t = sum(row["target"].values())
    deck_a = sum(row["actual"].values())
    if F != 0 and E == F:                                  return "In Hand"
    if E != 0 and E < F and deck_t == 0 and deck_a > 0:     return "Extra-Sub"
    if E == 0 and E < F and deck_t == 0 and deck_a > 0:     return "Bench-Sub"
    if E == 0 and F > 0:                                    return "Bench"
    if E < F:                                               return "Extra"
    if J == 0 and E > F:                                    return "To Buy"
    if J + F < E:                                           return "To Buy"
    if J + F >= E and J != 0:                               return "Ordered"
    if E == 0 and F == 0 and D == "B3":                     return "B3 Option"
    if E == 0 and F == 0:                                   return "Remove"
    return "To Buy"


CARD_SHEET = "Flat, no formulas"   # resolved values
FORMULA_SHEET = "Master"           # same rows, formulas only


def sheet_rows(wb, title, header_row):
    ws = wb[title]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = [("" if c is None else str(c).strip()) for c in rows[header_row - 1]]
    out = []
    for r in rows[header_row:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({h: (r[i] if i < len(r) else None) for i, h in enumerate(hdr) if h})
    return out


def read_master(wb, sheet, header_row):
    cards = []
    for r in sheet_rows(wb, sheet, header_row):
        name = r.get("Card")
        if blank(name):
            continue
        target = {d: int(num(r.get(d + "-T"))) for d in IDS}
        actual = {d: int(num(r.get(d + "-A"))) for d in IDS}
        ordered = int(num(r.get("Ordered")))
        # B-T =IF(AND(J<>"",SUM(M:R)=0),J,"") -- an ordered copy with no deck
        # asking for it is a copy the bench is expecting.
        bench_target = ordered if (ordered and sum(target.values()) == 0) else 0
        bench_actual = int(num(r.get("B-A")))
        row = {
            "name": str(name).strip(),
            "bracket": (str(r.get("Bracket")).strip() if not blank(r.get("Bracket")) else ""),
            "target": target, "actual": actual,
            "benchTarget": bench_target, "benchActual": bench_actual,
            "qty": sum(target.values()) + bench_target,          # E =SUM(M:S)
            "own": sum(actual.values()) + bench_actual,          # F =SUM(T:Z)
            # "In Cart" is a vendor tag, not a count: TCG, "TCG x2", WF.
            "cartVendor": (str(r.get("In Cart")).strip() if not blank(r.get("In Cart")) else ""),
            "ordered": ordered,
            "price": (round(float(r["$ Each"]), 2) if not blank(r.get("$ Each")) else None),
            "type": (str(r.get("Type")).strip() if not blank(r.get("Type")) else ""),
            "subType": (str(r.get("Sub-Type")).strip() if not blank(r.get("Sub-Type")) else ""),
            "color": (str(r.get("Color")).strip() if not blank(r.get("Color")) else ""),
            "mv": (num(r.get("MV")) if not blank(r.get("MV")) else None),
            "series": (str(r.get("Series")).strip() if not blank(r.get("Series")) else ""),
            "purpose": (str(r.get("Primary Purpose")).strip() if not blank(r.get("Primary Purpose")) else ""),
            "mechanics": [str(r.get(k)).strip() for k in
                          ("Primary Mechanic", "Secondary Mechanic", "Tertiary Mechanic")
                          if not blank(r.get(k))],
            "notes": (str(r.get("Notes")).strip() if not blank(r.get("Notes")) else ""),
            "moves": (str(r.get("Moves")).strip() if not blank(r.get("Moves")) else ""),
        }
        row["bench"] = max(0, row["own"] - row["qty"])           # G
        row["buyCount"] = max(0, row["qty"] - row["own"] - row["ordered"])   # H
        row["toBuyCost"] = (round(row["buyCount"] * row["price"], 2)
                            if row["buyCount"] > 0 and row["price"] is not None else 0.0)  # L
        row["status"] = status_of(row)
        # Whatever the sheet already resolved, so the caller can compare.
        row["_sheet"] = {"status": (str(r.get("Status")).strip() if not blank(r.get("Status")) else None),
                         "qty": r.get("Qty"), "own": r.get("Own"),
                         "bench": r.get("Extra (Bench)"), "buyCount": r.get("Buy Count"),
                         "toBuyCost": r.get("$ To Buy")}
        cards.append(row)
    return cards


def cross_check(cards):
    """The flat sheet's own numbers against the ones the formulas produce."""
    fields = [("status", str), ("qty", num), ("own", num),
              ("bench", num), ("buyCount", num), ("toBuyCost", num)]
    checked, bad = 0, []
    for c in cards:
        given = c.pop("_sheet")
        if given["status"] is None:
            continue                      # formula-only workbook: nothing to compare
        checked += 1
        for key, cast in fields:
            want = given[key]
            if want is None and key != "status":
                want = 0
            got = c[key]
            if cast is str:
                if str(want).strip() != got:
                    bad.append((c["name"], key, want, got))
            elif abs(num(want) - float(got)) > 0.005:
                bad.append((c["name"], key, want, got))
    return checked, bad


# "D5 Tuned add; targeted 2026-09-02; replaces Great Fierce Bee" -- the clauses
# in between vary and may themselves be semicolon-separated, so the gap is
# non-greedy across anything that is not another deck marker.
PAIR_RE = re.compile(r"D(\d) Tuned add(?:(?!D\d Tuned add).)*?replaces ([^;(]+)")


def tuned_pairs(cards):
    """The Master sheet's Notes column states each Tuned swap as a sentence:
    "D5 Tuned add; targeted 2026-09-02; replaces Great Fierce Bee". That is the
    authoritative record -- it carries later edit dates than the Upgrades sheet,
    and it survived the paste that corrupted two rows there."""
    pairs = []
    for c in cards:
        for deck_no, replaces in PAIR_RE.findall(c["notes"] or ""):
            pairs.append({"deck": "D" + deck_no, "add": c["name"],
                          "replaces": replaces.strip(), "price": c["price"]})
    return pairs


def read_upgrades(wb, cards, master_by_name):
    """The D5/D6 Tuned swaps: what goes in, and what comes out to the bench.

    Two records of the same plan disagree. The Upgrades sheet was built
    2026-08-31; the Master sheet's Notes carry pairings dated 2026-09-02 and name
    four D5 cuts the Upgrades sheet never got. Rather than union two records that
    contradict each other, the Upgrades sheet stays the cut list, the Notes
    supply the "replaces" line on each add, and the disagreement is reported in
    `dataNotes` instead of being smoothed away.

    One ADD row lost both its Deck and Card cells to a bad paste and is put back
    from its price; one CUT row lost its Card cell and cannot be identified --
    four candidates fit -- so it is dropped and flagged rather than guessed.
    """
    rows = [r for r in sheet_rows(wb, "Upgrades (D5-D6)", 1)
            if str(r.get("Action") or "").strip() in ("ADD", "CUT")]
    pairs = [p for p in tuned_pairs(cards) if p["deck"] in ("D5", "D6")]
    by_add = {p["add"]: p for p in pairs}

    out, notes, seen_add = [], [], set()
    for r in rows:
        action = str(r["Action"]).strip()
        deck = str(r.get("Deck") or "").strip()
        card = str(r.get("Card") or "").strip()
        repaired = ""
        if card not in master_by_name:
            price = r.get("$ Each")
            hits = [p for p in pairs if p["price"] is not None and not blank(price)
                    and abs(p["price"] - float(price)) < 0.005] if action == "ADD" else []
            if len(hits) == 1:
                deck, card = hits[0]["deck"], hits[0]["add"]
                repaired = "name and deck recovered from its price and the Notes pairing"
            else:
                notes.append("Upgrades sheet: a {0} row for {1} lost its card name to a paste "
                             "and cannot be identified; it is left out of the counts."
                             .format(action, deck or "an unknown deck"))
                continue
        if deck not in IDS:
            continue
        if action == "ADD":
            seen_add.add(card)
        out.append({"Deck": deck, "Action": action, "Card": card,
                    "$ Each": r.get("$ Each") if not blank(r.get("$ Each"))
                              else (master_by_name.get(card) or {}).get("price"),
                    "Status": r.get("Status"), "Notes": r.get("Notes"),
                    "Replaces": by_add.get(card, {}).get("replaces", "") if action == "ADD" else "",
                    "_repaired": repaired})

    # An add the sheet never got is a card that would otherwise vanish from the
    # deck's upgrade list, so those are put back.
    for p in pairs:
        if p["add"] in seen_add:
            continue
        out.append({"Deck": p["deck"], "Action": "ADD", "Card": p["add"], "$ Each": p["price"],
                    "Status": (master_by_name.get(p["add"]) or {}).get("status"),
                    "Notes": "", "Replaces": p["replaces"],
                    "_repaired": "absent from the Upgrades sheet; taken from the Notes pairing"})

    cut_names = {r["Card"] for r in out if r["Action"] == "CUT"}
    stray = sorted({p["replaces"] for p in pairs
                    if p["replaces"] not in cut_names and p["replaces"] in master_by_name})
    if stray:
        notes.append("The Master Notes (2026-09-02) name {0} cards as replaced that the Upgrades "
                     "sheet (2026-08-31) does not list as cuts: {1}. The Notes are newer; the cut "
                     "counts below follow the sheet.".format(len(stray), ", ".join(stray)))
    return out, notes


def price_fallback(root):
    """Scryfall prices for the rows the workbook leaves blank.

    Seventy bench rows carry no "$ Each" -- they are cards nobody costed because
    nobody was buying them -- which left the Bench tab showing a dash where a
    price belongs. data/card-facts.json already holds a Scryfall price for every
    one of them.

    That file is generated *from* this one, so the order is: import, build the
    facts, import again. The second pass is what fills the prices, and it is
    idempotent -- a row that already has a workbook price is never overwritten,
    because the workbook is what he actually paid.
    """
    path = os.path.join(root, "data", "card-facts.json")
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as fh:
        return {name: fact.get("price") for name, fact in
                json.load(fh).get("cards", {}).items() if fact.get("price") is not None}


def main():
    default = os.path.join(ROOT, "data", "source", "MtG_Deck_Flat.xlsx")
    src = sys.argv[1] if len(sys.argv) > 1 else default
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    if CARD_SHEET in wb.sheetnames:
        cards = read_master(wb, CARD_SHEET, 1)
    else:
        cards = read_master(wb, FORMULA_SHEET, 4)
    checked, bad = cross_check(cards)
    if bad:
        for row in bad[:20]:
            print("  MISMATCH {0}  {1}: sheet={2!r} formula={3!r}".format(*row), file=sys.stderr)
        sys.exit("{0} cells disagree between the flat sheet and the formulas".format(len(bad)))
    print("cross-check: {0} rows, flat sheet and formulas agree".format(checked) if checked
          else "cross-check: workbook carries no resolved values; using the formulas")
    fallback = price_fallback(ROOT)
    filled = 0
    for card in cards:
        card["priceSource"] = "workbook" if card["price"] is not None else ""
        if card["price"] is None and card["name"] in fallback:
            card["price"] = round(float(fallback[card["name"]]), 2)
            card["priceSource"] = "scryfall"
            filled += 1
            if card["buyCount"] > 0:
                card["toBuyCost"] = round(card["buyCount"] * card["price"], 2)
    if filled:
        print("filled {0} blank prices from Scryfall".format(filled))

    by_name = {c["name"]: c for c in cards}

    upgrades, upgrade_notes = read_upgrades(wb, cards, by_name)
    b3 = [r for r in sheet_rows(wb, "B3 Upgrades", 1)
          if str(r.get("Deck") or "").strip() in IDS]

    # The Read Me names each deck's commander; the Master tags one card in each
    # deck with the purpose "Commander". Where the two disagree the deck may be
    # illegal, so the disagreement is reported rather than resolved silently.
    for did, label, commander in DECKS:
        tagged = [c["name"] for c in cards
                  if c["target"][did] > 0 and c["purpose"] == "Commander"]
        if tagged and commander not in tagged:
            upgrade_notes.append(
                "{0}: the Read Me names {1} as the commander, but the Master tags {2} "
                "with the Commander purpose. Only a legendary creature, or a planeswalker "
                "whose text says it can be your commander, is legal in the slot; the Read "
                "Me's card is used here.".format(did, commander, " and ".join(tagged)))

    decks = []
    for did, label, commander in DECKS:
        picked = [c for c in cards if c["target"][did] > 0]
        in_box = [c for c in cards if c["actual"][did] > 0]
        decks.append({
            "id": did, "label": label, "commander": commander,
            "targetCards": sum(c["target"][did] for c in picked),
            "boxCards": sum(c["actual"][did] for c in in_box),
            "upgrades": [
                {"action": str(r["Action"]).strip(), "card": str(r.get("Card") or "").strip(),
                 "price": (round(float(r["$ Each"]), 2) if not blank(r.get("$ Each")) else None),
                 "status": str(r.get("Status") or "").strip(),
                 "note": str(r.get("Notes") or "").strip(),
                 "replaces": str(r.get("Replaces") or "").strip(),
                 "repaired": r.get("_repaired", "")}
                for r in upgrades if str(r["Deck"]).strip() == did],
            "b3": [
                {"add": str(r["B3 Add"]).strip(),
                 "replaces": str(r.get("Replaces (-> Bench)") or "").strip(),
                 "price": (round(float(r["$ Each"]), 2) if not blank(r.get("$ Each")) else None),
                 "gameChanger": not blank(r.get("GC")),
                 "color": str(r.get("Color") or "").strip(),
                 "mv": (num(r.get("MV")) if not blank(r.get("MV")) else None),
                 "type": str(r.get("Type") or "").strip(),
                 "status": str(r.get("Status") or "").strip(),
                 "why": str(r.get("Why It Makes B3") or "").strip()}
                for r in b3 if str(r["Deck"]).strip() == did],
        })

    out = {
        "schemaVersion": 1,
        "generatedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "source": os.path.basename(src),
        "crossChecked": checked,
        "note": ("Formula columns (Status, Qty, Own, Extra, Buy Count, $ To Buy) are "
                 "recomputed here: the workbook ships with no cached formula results."),
        "dataNotes": upgrade_notes,
        "decks": decks,
        "cards": cards,
    }
    dest = os.path.join(ROOT, "data", "master-v2.json")
    with open(dest, "w", encoding="utf-8") as fh:
        json.dump(out, fh, indent=1, ensure_ascii=False)
        fh.write("\n")
    print(f"wrote {dest}  ({len(cards)} cards, {len(decks)} decks)")
    for d in decks:
        print(f"  {d['id']} {d['label']:<18} target {d['targetCards']:>3}  in box {d['boxCards']:>3}"
              f"  upgrades {len(d['upgrades']):>2}  B3 {len(d['b3']):>2}")


if __name__ == "__main__":
    main()
